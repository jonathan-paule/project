import boto3
from openpyxl import Workbook
from openpyxl.styles import Font, numbers
import re
import os
from datetime import datetime
from email.mime.multipart import MIMEMultipart
from email.mime.application import MIMEApplication
from email.mime.text import MIMEText

# ---------- Validation Helpers ----------
def validate_email(email: str) -> str:
    pattern = r'^[A-Za-z0-9._%+-]+@adcuratio\.com$'
    if not re.match(pattern, email):
        raise ValueError("❌ Invalid email. Must be in format user@adcuratio.com")
    return email

def validate_date(date_text: str) -> str:
    try:
        datetime.strptime(date_text, "%Y-%m-%d")
        return date_text
    except ValueError:
        raise ValueError("❌ Invalid date format. Use YYYY-MM-DD")

def ensure_date_order(start_date: str, end_date: str) -> None:
    s = datetime.strptime(start_date, "%Y-%m-%d")
    e = datetime.strptime(end_date, "%Y-%m-%d")
    if e < s:
        raise ValueError("❌ End date cannot be before start date.")

# ---------- AWS Cost Export ----------
def export_costs_to_excel(start_date: str, end_date: str, filename: str = "aws_costs_calculator.xlsx") -> str:
    ce = boto3.client("ce", region_name="us-east-1")

    params = {
        "TimePeriod": {"Start": start_date, "End": end_date},
        "Granularity": "MONTHLY",
        "Metrics": ["BlendedCost"],
        "GroupBy": [
            {"Type": "DIMENSION", "Key": "SERVICE"},
            {"Type": "TAG", "Key": "Name"},
        ],
    }

    wb = Workbook()
    ws = wb.active
    ws.title = "AWS Costs"

    # Header row
    ws.append(["Service", "ResourceName(Tag:Name)", "Cost", "Unit"])
    header_font = Font(bold=True)
    for col in range(1, 5):
        ws.cell(row=1, column=col).font = header_font

    bold_font = Font(bold=True)
    current_service = None

    while True:
        response = ce.get_cost_and_usage(**params)

        for result in response.get("ResultsByTime", []):
            for group in result.get("Groups", []):
                service = group["Keys"][0]
                name_key = group["Keys"][1]
                tag_value = (name_key.replace("Name$", "") or "NoNameTag").strip()

                amount_str = group["Metrics"]["BlendedCost"]["Amount"]
                try:
                    amount = float(amount_str)
                except (TypeError, ValueError):
                    amount = 0.0
                unit = group["Metrics"]["BlendedCost"]["Unit"]

                if current_service != service:
                    # Add service row with cost 0
                    ws.append([f"Service: {service}", "", 0, "USD"])
                    ws.cell(row=ws.max_row, column=1).font = bold_font
                    ws.cell(row=ws.max_row, column=3).number_format = numbers.FORMAT_NUMBER_00
                    current_service = service

                # Add the resource row with number formatting for cost
                ws.append([service, tag_value, amount, unit])
                ws.cell(row=ws.max_row, column=3).number_format = numbers.FORMAT_NUMBER_00

        token = response.get("NextPageToken")
        if token:
            params["NextPageToken"] = token
        else:
            break

    # Add TOTAL row at the bottom
    ws.append([])
    last_data_row = ws.max_row
    ws.append(["TOTAL", "", f"=SUM(C2:C{last_data_row})", "USD"])
    ws.cell(row=ws.max_row, column=1).font = bold_font
    ws.cell(row=ws.max_row, column=3).font = bold_font
    ws.cell(row=ws.max_row, column=3).number_format = numbers.FORMAT_NUMBER_00

    wb.save(filename)
    print(f"✅ Costs exported to {filename}")
    return filename

# ---------- Email Sending (SES) ----------
def send_email(recipient_email: str, subject: str, body: str, attachment_path: str):
    ses = boto3.client("ses", region_name="us-east-1")

    sender = "your_verified_email@adcuratio.com"  # must be verified in SES

    # Build email
    msg = MIMEMultipart()
    msg["Subject"] = subject
    msg["From"] = sender
    msg["To"] = recipient_email

    # Body
    msg.attach(MIMEText(body, "plain"))

    # Attachment
    with open(attachment_path, "rb") as f:
        part = MIMEApplication(f.read())
        part.add_header(
            "Content-Disposition",
            f"attachment; filename={os.path.basename(attachment_path)}"
        )
        msg.attach(part)

    # Send email via SES
    response = ses.send_raw_email(
        Source=sender,
        Destinations=[recipient_email],
        RawMessage={"Data": msg.as_string()}
    )

    print(f"📧 Email sent to {recipient_email}, MessageId: {response['MessageId']}")

# ---------- Main ----------
if __name__ == "__main__":
    try:
        start_date = validate_date(input("Enter start date (YYYY-MM-DD): ").strip())
        end_date = validate_date(input("Enter end date (YYYY-MM-DD): ").strip())
        ensure_date_order(start_date, end_date)

        recipient_email = validate_email(input("Enter recipient email (must end with @adcuratio.com): ").strip())

        filename = export_costs_to_excel(start_date, end_date)

        subject = f"AWS Cost Report {start_date} to {end_date}"
        body = f"Please find attached the AWS cost report from {start_date} to {end_date}."
        send_email(recipient_email, subject, body, filename)

    except ValueError as ve:
        print(ve)
    except FileNotFoundError as fe:
        print(fe)
    except Exception as e:
        print(f"⚠️ Unexpected error: {e}")
